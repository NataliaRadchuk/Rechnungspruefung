import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill
import chardet
import locale
from ui.gui_log_handler import setup_logger
#import logging

class FileHandler:
    def __init__(self, log_callback):
        #logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        locale.setlocale(locale.LC_NUMERIC, 'de_DE.UTF-8')  # Setze das deutsche Zahlenformat
        self.logger = setup_logger(log_callback) # Initializiere das Logging

    def detect_encoding(self, file_path):
        with open(file_path, 'rb') as f:
            result = chardet.detect(f.read(100000))
        return result['encoding']

    def read_file(self, file_path):
        file_extension = os.path.splitext(file_path)[1].lower()
        try:
            if file_extension == '.csv':
                encoding = self.detect_encoding(file_path)
                self.logger.info(f"Detected file encoding: {encoding}")
                
                if encoding.lower() in ['utf-8', 'utf-8-sig', 'utf-16', 'utf-16-le', 'utf-16-be']:
                    if encoding.lower() in ['utf-8-sig', 'utf-8']:
                        df = pd.read_csv(file_path, encoding='utf-8-sig', sep=None, engine='python')
                    elif encoding.lower().startswith('utf-16'):
                        df = pd.read_csv(file_path, encoding='utf-16', sep=None, engine='python')
                else:
                    raise ValueError(f"Unsupported file encoding: {encoding}")
                
                self.logger.info(f"DataFrame head:\n{df.head()}")
                return df
            else:
                raise ValueError("Unsupported file extension")
        except Exception as e:
            logging.error(f"Error reading file: {e}")
            return None

    def copy_and_trim_file(self, input_file):
        try:
            df = self.read_file(input_file)
            if df is None or df.empty:
                raise ValueError("DataFrame is empty or file could not be read.")
            
            # Nur die erste Spalte entfernen, aber alle Datenzeilen behalten
            trimmed_df = df.iloc[:, 1:]
            self.logger.info(f"Trimmed DataFrame:\n{trimmed_df.head()}")
            return trimmed_df
        except Exception as e:
            logging.error(f"An error occurred: {e}")
            return None

    def process_cell(self, ws, row, col, value, col_letter, integer_columns, number_columns):
        if col_letter in integer_columns and pd.notna(value):
            try:
                integer_value = int(float(str(value).replace(',', '.')))
                cell = ws.cell(row=row, column=col + 1, value=integer_value)
                cell.number_format = '0'
            except ValueError:
                logging.warning(f"Could not convert '{value}' to integer in column {col_letter}. Inserting as text.")
                cell = ws.cell(row=row, column=col + 1, value=value)
        elif col_letter in number_columns and pd.notna(value):
            try:
                numeric_string = str(value).replace('.', '').replace(',', '.')
                numeric_value = float(numeric_string)
                cell = ws.cell(row=row, column=col + 1, value=numeric_value)
                cell.number_format = '#,##0.00'
            except ValueError:
                logging.warning(f"Could not convert '{value}' to float. Inserting as text.")
                cell = ws.cell(row=row, column=col + 1, value=value)
        else:
            cell = ws.cell(row=row, column=col + 1, value=value)

    def append_trimmed_to_existing(self, trimmed_df, preset_file):
        try:
            if trimmed_df is None or trimmed_df.empty:
                raise ValueError("Trimmed DataFrame is empty.")

            wb = load_workbook(preset_file, keep_vba=True)
            if "Prüf" not in wb.sheetnames:
                raise ValueError('Worksheet "Prüf" not found in the preset file.')

            ws = wb["Prüf"]

            start_row = next((row for row, cells in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2) if all(cell is None for cell in cells)), ws.max_row + 1)

            number_columns = {'F', 'G', 'H', 'I', 'J'}
            integer_columns = {'A', 'E'}

            # Verarbeitung der Hauptdaten
            for r_idx, row in enumerate(trimmed_df.itertuples(index=False), start=start_row):
                for c_idx, value in enumerate(row):
                    col_letter = chr(65 + c_idx)
                    self.process_cell(ws, r_idx, c_idx, value, col_letter, integer_columns, number_columns)

            last_row = start_row + len(trimmed_df) - 1

            # Verschieben der Zusammenfassungszeile
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=last_row + 1, column=col_idx, value=ws.cell(row=last_row, column=col_idx).value)
                ws.cell(row=last_row, column=col_idx).value = None

            # Entfernen der "Gesamtwert"-Einträge
            for col_idx in range(1, 4):
                ws.cell(row=last_row + 1, column=col_idx).value = ""

            # Hinzufügen eines schwarzen Strichs über der letzten Zeile
            black_border = Border(top=Side(style='thin', color='000000'))
            for col in ws[last_row]:
                col.border = black_border

            # Neuer Code: Einfügen des spezifischen Inhalts
            last_row += 3  # Zwei Leerzeilen nach der letzten Zeile

            # "gebucht" in Spalte G
            ws.cell(row=last_row, column=7, value="gebucht")

            # Kopieren des Inhalts der letzten Zeile in Spalte H
            ws.cell(row=last_row, column=8, value=ws.cell(row=last_row-2, column=8).value)

            # Hinzufügen der Wörter in Spalte J und Berechnungen in Spalte I
            words = ["Hotelrechnung", "Rabatt", "Rechnung", "Depo", "Guthaben"]
            for i, word in enumerate(words):
                ws.cell(row=last_row+i, column=10, value=word)

            # Berechnung für "Rechnung" in Spalte I
            ws.cell(row=last_row+2, column=9, value=f"=SUM(I{last_row}:I{last_row+1})")

            # Berechnung für "Guthaben" in Spalte I
            ws.cell(row=last_row+4, column=9, value=f"=SUM(I{last_row+2}:I{last_row+3})")

            # Striche und Formatierungen
            for col in range(9, 11):
                ws.cell(row=last_row+2, column=col).border = Border(top=Side(style='thin'))
                ws.cell(row=last_row+4, column=col).border = Border(top=Side(style='thin'))
                ws.cell(row=last_row+5, column=col).border = Border(top=Side(style='double'))

            # "Rechnung" und "Guthaben" fett machen
            ws.cell(row=last_row+2, column=10).font = Font(bold=True)
            ws.cell(row=last_row+4, column=10).font = Font(bold=True)

            # "Diff" in Spalte G statt H, fett und neon gelb markiert
            diff_cell = ws.cell(row=last_row+4, column=7, value="Diff")
            diff_cell.font = Font(bold=True)
            for col in range(7, 9):
                ws.cell(row=last_row+4, column=col).fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
            # Neue Berechnung neben "Diff"
            ws.cell(row=last_row+4, column=8, value=f"=I{last_row}-H{last_row}")

            self.logger.info(f"Successfully appended trimmed data to worksheet 'Prüf'")
            return wb
        except Exception as e:
            logging.error(f"An error occurred: {e}")
            return None

    def process_report(self, input_report):
        try:
            trimmed_df = self.copy_and_trim_file(input_report)
            if trimmed_df is None:
                raise ValueError("Failed to trim input report.")
            
            # Set the correct path to the template file
            current_dir = os.path.dirname(os.path.abspath(__file__))
            preset_file = os.path.join(current_dir, "templates", "template.xlsm")
            
            template_filled = self.append_trimmed_to_existing(trimmed_df, preset_file)
            if template_filled is None:
                raise ValueError("Failed to append trimmed data to template.")
            
            return template_filled
        except Exception as e:
            logging.error(f"Error in process_report: {e}")
            return None

# Verwendung der FileHandler-Klasse
if __name__ == "__main__":
    file_handler = FileHandler()
    input_file = 'path_to_input_file.csv'
    template_filled = file_handler.process_report(input_file)
    if template_filled:
        print("Report processed successfully.")
    else:
        print("Failed to process report.")