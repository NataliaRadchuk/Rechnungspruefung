import pandas as pd
import os
import logging
import chardet
from datetime import datetime
import locale
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from file_handler import FileHandler 

class NameListHandler:
    def __init__(self):
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        locale.setlocale(locale.LC_NUMERIC, 'de_DE.UTF-8')  # Setze das deutsche Zahlenformat
        self.file_ops = FileHandler()

    def prepare_table(self, name_list_file):
        pandas_file = self.file_ops.copy_and_trim_file(name_list_file)
    
        if pandas_file is None or pandas_file.empty:
            logging.error("Failed to load or process the file, or the file is empty.")
            return None

        logging.info("File loaded and trimmed successfully.")
        logging.info(f"Trimmed DataFrame:\n{pandas_file}")

        # Aufgabe 1: Letzte Spalte löschen, wenn sie nur leere Werte enthält
        if pandas_file.iloc[:, -1].astype(str).str.strip().eq('').all():
            pandas_file = pandas_file.iloc[:, :-1]

        # Aufgabe 2: Spalte K (Roomnights) löschen, falls vorhanden
        if pandas_file.shape[1] >= 11:
            pandas_file = pandas_file.iloc[:, :10]

        # Aufgabe 4: Spalte A löschen (falls noch nicht geschehen)
        if pandas_file.shape[1] > 1:
            pandas_file = pandas_file.iloc[:, 1:]

        # Aufgabe 3: Inhalt von Spalte I nach J verschieben, I leer lassen
        if pandas_file.shape[1] >= 9:
            # Füge eine neue leere Spalte am Ende hinzu
            pandas_file['New'] = ''
            # Verschiebe den Inhalt von I nach J
            pandas_file.iloc[:, -1] = pandas_file.iloc[:, -2]
            # Leere Spalte I
            pandas_file.iloc[:, -2] = ''
        
        # Stelle sicher, dass wir genau 10 Spalten haben
        if pandas_file.shape[1] < 10:
            for i in range(pandas_file.shape[1], 10):
                pandas_file[f'New_{i}'] = ''
        elif pandas_file.shape[1] > 10:
            pandas_file = pandas_file.iloc[:, :10]

        # Spalten neu indizieren
        pandas_file.columns = [chr(65 + i) for i in range(10)]

        # Aufgabe 7: Zeilen basierend auf Werten in Spalte H duplizieren
        new_rows = []
        for index, row in pandas_file.iterrows():
            h_value = row['H']
            try:
                h_value = int(float(h_value))  # Versuche, den Wert in eine Ganzzahl umzuwandeln
                if h_value > 0:
                    for i in range(h_value):
                        new_row = row.copy()
                        new_row['H'] = 1
                        new_rows.append(new_row)
                else:
                    new_rows.append(row)  # Behalte die Originalzeile für Werte <= 0
            except ValueError:
                new_rows.append(row)  # Behalte die Originalzeile, wenn H keine Zahl ist

        pandas_file = pd.DataFrame(new_rows, columns=pandas_file.columns)

        # Aufgabe 2: Kopiere und füge den gesamten Inhalt ein, füge "-F" in Spalte F hinzu
        copied_df = pandas_file.copy()
        copied_df['F'] = copied_df['F'].astype(str) + '-F'
        pandas_file = pd.concat([pandas_file, copied_df], ignore_index=True)

        # Alphabetisch sortieren nach Spalte C
        pandas_file = pandas_file.sort_values(by='C')

        logging.info(f"Prepared DataFrame:\n{pandas_file}")
        return pandas_file
    
    def append_into_preset(self, trimmed_df, workbook):
        logging.info("In append to preset")

        # Wähle den Tab "SR Lt. AM"
        sheet = workbook['SR Lt. AM']

        # Aufgabe A: Füge trimmed_df ab Zeile 18, Spalte A ein
        last_data_row = 17  # Initialisiere mit der Zeile vor den Daten
        for r, row in enumerate(dataframe_to_rows(trimmed_df, index=False, header=False), start=18):
            # Fülle die ersten drei Spalten (A, B, C)
            for c in range(3):
                cell = sheet.cell(row=r, column=c+1)
                value = row[c] if c < len(row) else ''
                if isinstance(value, (int, float)):
                    cell.value = value
                else:
                    cell.value = str(value)

            for c in range(3, len(row)):
                cell = sheet.cell(row=r, column=c+6)  # +6 weil wir 5 Spalten übersprungen haben
                value = row[c]
                if isinstance(value, (int, float)):
                    cell.value = value
                else:
                    cell.value = str(value)
                    
            # Füge die Formel in Spalte N ein (ehemals Spalte I)
            sheet.cell(row=r, column=14).value = f'=J{r}-I{r}'
            last_data_row = r  # Aktualisiere die letzte Datenzeile

        # Konvertiere Spalte I und J in Datumsformat (ehemals D und E)
        def convert_to_date(cell):
            if cell.value:
                try:
                    date_value = datetime.strptime(str(cell.value), "%d.%m.%Y")
                    cell.value = date_value
                    cell.number_format = 'DD.MM.YYYY'
                except ValueError:
                    logging.warning(f"Konnte Datum in Zelle {cell.coordinate} nicht konvertieren: {cell.value}")

        for col in [9, 10]:  # 9 für Spalte I, 10 für Spalte J
            for row in range(18, last_data_row + 1):
                convert_to_date(sheet.cell(row=row, column=col))
        
        # Aufgabe 3: Einträge aus Spalte F in die innere Tabelle einfügen
        unique_entries = sorted(set(trimmed_df['F'].astype(str)))
        
        for i, entry in enumerate(unique_entries):
            if i < 6:  # Für die ersten 6 Einträge
                sheet.cell(row=i+3, column=16).value = entry  # Spalte P ist die neue Spalte K
        
        # Suche nach "Gesamt" in Spalte P ab Zeile 10 (ehemals Spalte K)
        for row in range(10, sheet.max_row + 1):
            if sheet.cell(row=row, column=16).value == "Gesamt":
                # Passe die Formel an, um nur den Bereich mit Daten zu berücksichtigen
                sheet.cell(row=row, column=17).value = f'=MIN(I18:I{last_data_row})-2'  # Spalte I ist die neue Spalte D
                sheet.cell(row=row, column=17).number_format = 'DD.MM.YYYY'
                break

        # Aufgabe 1: Füge "Ist" ein
        ist_row = last_data_row + 2
        sheet.cell(row=ist_row, column=15).value = "Ist"
        
        # Aufgabe 2: Berechne Summe/2 für Spalte N
        sheet.cell(row=ist_row, column=14).value = f'=SUM(N18:N{last_data_row})/2'
        
        # Aufgabe 3: Berechne Summe für Spalte P
        sheet.cell(row=ist_row, column=16).value = f'=SUM(P18:P{last_data_row})'
        
        # Aufgabe 4: Füge "Soll" ein
        soll_row = ist_row + 2
        sheet.cell(row=soll_row, column=15).value = "Soll"
        
        # Aufgabe 5: Übernehme Wert aus Tab "Prüf" Spalte E
        sheet.cell(row=soll_row, column=14).value = '=INDEX(Prüf!E:E, MATCH(1E+99, Prüf!E:E))'
        
        # Aufgabe 6: Übernehme Wert aus Tab "Prüf" Spalte H
        sheet.cell(row=soll_row, column=16).value = '=INDEX(Prüf!H:H, MATCH(1E+99, Prüf!H:H) - 6)'
        
        # Aufgabe 7: Füge "Diff" ein
        diff_row = soll_row + 2
        sheet.cell(row=diff_row, column=15).value = "Diff"
        
        # Aufgabe 8: Berechne Differenz für Spalte N
        sheet.cell(row=diff_row, column=14).value = f'=N{ist_row}-N{soll_row}'
        
        # Aufgabe 9: Berechne Differenz für Spalte P
        sheet.cell(row=diff_row, column=16).value = f'=P{ist_row}-P{soll_row}'
    
        return workbook
    
    def process_template(self, template_filled, input_namelist, checkset):
        try:
            prepared_table = self.prepare_table(input_namelist)
            if prepared_table is None:
                raise ValueError("Failed to prepare name list table.")
            
            fully_filled_template = self.append_into_preset(prepared_table, template_filled)
            if fully_filled_template is None:
                raise ValueError("Failed to append prepared table into template.")
            
            return fully_filled_template
        except Exception as e:
            logging.error(f"Error in process_template: {e}")
            return None

# Verwendung der NameListHandler-Klasse
if __name__ == "__main__":
    name_list_handler = NameListHandler()
    input_file = 'path_to_input_file.csv'
    template_filled = openpyxl.load_workbook('path_to_template_filled.xlsm', keep_vba=True)
    
    fully_filled_template = name_list_handler.process_template(template_filled, input_file, True)
    if fully_filled_template:
        print("Template processed successfully.")
        fully_filled_template.save('path_to_output_file.xlsm')
    else:
        print("Failed to process template.")